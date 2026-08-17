from __future__ import annotations
import os
import re
from heapq import nsmallest
from math import log2
from itertools import combinations, product


def nafTerms(x: int) -> list[tuple[int, int]]:
    '''Return the list of (sign, k) pairs for the nonzero terms in the NAF representation of x, where sign is either 1 or -1, and k is the exponent of 2'''
    if x == 0:
        return []
    naf = []
    outerSign = 1 if x > 0 else -1
    x = abs(x)
    k = 0
    while x > 0:
        if x & 1:
            u = 2 - (x & 3)
            naf.append((outerSign * u, k))
            x -= u
        x //= 2
        k += 1
    return naf
            

def nafTermsCount(x: int) -> int:
    '''Return the number of nonzero terms in the NAF representation of x'''
    return len(nafTerms(x))


def nafTermsMaxPower(x: int) -> int:
    '''Return the maximum power of 2 among the nonzero terms in the NAF representation of x, or -1 if x is 0'''
    terms = nafTerms(x)
    if not terms:
        return -1
    return max(k for _, k in terms)


def nafTermsModulusLiftBeamSearch(x: int, modulus: int, maxPower: int, maxMultipleOfModulus: int, maxSearchDepth: int, beamWidth: int) -> tuple[int, list[tuple[int, int]], int, int]:
    '''Return the integer y such that y ≡ x (mod modulus) and the NAF representation of y has the smallest number of nonzero terms among all integers congruent to x modulo modulus and with maximum power of 2 among the nonzero terms in its NAF representation at most maxPower, along with the list of (sign, k) pairs for the nonzero terms in the NAF representation of y, the number of nonzero terms, and the maximum power of 2 among the nonzero terms'''
    def score(x: int) -> tuple[int, int, int]:
        termsCount = nafTermsCount(x)
        termsMaxPower = nafTermsMaxPower(x)
        if termsMaxPower > maxPower:
            return (10**18, 10**18, 10**18)
        absValue = abs(x)
        return (termsCount, absValue, termsMaxPower)
    
    r = x % modulus
    frontier = [r, r - modulus]
    best = min(frontier, key=score)
    if maxMultipleOfModulus < 1:
        raise ValueError(f'maxMultipleOfModulus {maxMultipleOfModulus} must be at least 1')
    maxShift = int(log2(maxMultipleOfModulus)) + 1

    for _ in range(maxSearchDepth):
        candidates = []
        for c in frontier:
            for shift in range(maxShift + 1):
                step = (modulus << shift)
                candidates.append(c + step)
                candidates.append(c - step)
        frontier = nsmallest(beamWidth, candidates, key=score)
        bestCandidate = min(frontier, key=score)
        if score(bestCandidate) < score(best):
            best = bestCandidate

    return best, nafTerms(best), nafTermsCount(best), nafTermsMaxPower(best)


def nafTermsModulusLiftTargetFirstSearch(x: int, modulus: int, maxPower: int, maxNumberOfTerms: int) -> tuple[int, list[tuple[int, int]], int, int] | None:
    '''Find y ≡ x (mod modulus) whose NAF has the fewest nonzero terms (≤ maxNumberOfTerms) with every term's power ≤ maxPower. Returns (y, nafTerms(y), termsCount, maxExponent). Return None if no such y exists in range.'''
    if maxNumberOfTerms < 1:
        raise ValueError(f'maxNumberOfTerms {maxNumberOfTerms} must be at least 1')
    if maxPower < 1:
        raise ValueError(f'maxPower {maxPower} must be at least 1')
    
    r = x % modulus
    if r == 0:
        return 0, [], 0, -1
    
    n = maxPower + 1
    for k in range(1, maxNumberOfTerms + 1):
        # check if k combinations can fit with current maxPower
        if 2 * k - 1 > n:
            break
        for base in combinations(range(n - k + 1), k):
            powers = tuple(p + i for i, p in enumerate(base))
            for signs in product([-1, 1], repeat=k):
                y = sum(s << p for s, p in zip(signs, powers))
                if y % modulus == r:
                    return y, list(zip(signs, powers)), k, powers[-1]
    
    return None


def nafTermsModulusLift(x: int, modulus: int, maxPower: int, maxMultipleOfModulus: int, maxSearchDepth:int, beamWidth: int, maxNumberOfTerms: int) -> tuple[int, list[tuple[int, int]], int, int]:
    '''Return the integer y such that y ≡ x (mod modulus) and the NAF representation of y has the smallest number of nonzero terms among all integers congruent to x modulo modulus and with maximum power of 2 among the nonzero terms in its NAF representation at most maxPower, along with the list of (sign, k) pairs for the nonzero terms in the NAF representation of y, the number of nonzero terms, and the maximum power of 2 among the nonzero terms. This function will first try to find such integer using target-first search, and if it fails, it will fall back to beam search'''
    targetFirstResult = nafTermsModulusLiftTargetFirstSearch(x, modulus, maxPower, maxNumberOfTerms)
    if targetFirstResult is not None:
        return targetFirstResult
    return nafTermsModulusLiftBeamSearch(x, modulus, maxPower, maxMultipleOfModulus, maxSearchDepth, beamWidth)


def bitReverse(x: int, bits: int) -> int:
    '''Reverse the low `bits` bits of x.'''
    result = 0
    for _ in range(bits):
        result = (result << 1) | (x & 1)
        x >>= 1
    return result


def vectorAdd(a: list[int], b: list[int]) -> list[int]:
    '''Element-wise addition. Mirrors IntType.__add__ on a value-batch.'''
    if len(a) != len(b):
        raise ValueError(f'vectorAdd: length mismatch, got {len(a)} and {len(b)}')
    return [x + y for x, y in zip(a, b)]


def vectorSub(a: list[int], b: list[int]) -> list[int]:
    '''Element-wise subtraction. Mirrors IntType.__sub__ on a value-batch.'''
    if len(a) != len(b):
        raise ValueError(f'vectorSub: length mismatch, got {len(a)} and {len(b)}')
    return [x - y for x, y in zip(a, b)]


def vectorMul(a: list[int], scalar: int) -> list[int]:
    '''Per-element multiplication by a scalar. Mirrors IntType.__mul__ with an int operand.'''
    return [x * scalar for x in a]


def vectorLshift(a: list[int], shift: int) -> list[int]:
    '''Per-element left shift. Mirrors IntType.__lshift__.'''
    if shift < 0:
        raise ValueError(f'vectorLshift: shift must be non-negative, got {shift}')
    return [x << shift for x in a]


def vectorRshift(a: list[int], shift: int) -> list[int]:
    '''Per-element right shift (Python arithmetic shift). Mirrors IntType.__rshift__.'''
    if shift < 0:
        raise ValueError(f'vectorRshift: shift must be non-negative, got {shift}')
    return [x >> shift for x in a]


def vectorSlice(a: list[int], start: int, end: int, signed: bool = False) -> list[int]:
    '''Per-element bit-slice over the inclusive range [start, end]. Mirrors IntType.slice: when signed=False (interior limb, width < shifted.bitWidth in the IntType case) returns the unsigned slice (x >> start) & mask. When signed=True (boundary limb, width >= shifted.bitWidth in the IntType case) returns the Python-signed (x >> start) without masking — preserving the sign extension that the Goldilocks limb-folding relies on for negative inputs.'''
    width = end - start + 1
    if width <= 0:
        raise ValueError(f'vectorSlice: invalid range ({start}, {end})')
    if signed:
        return [x >> start for x in a]
    mask = (1 << width) - 1
    return [(x >> start) & mask for x in a]


def vectorBitWidth(a: list[int]) -> int:
    '''Worst-case bit width across a value batch, mirroring IntType.bitWidth: 0 for an all-zero batch; max(x.bit_length()) for an all-non-negative batch; max(negWidth, posWidth) + 1 for a mixed/signed batch (the +1 is the sign bit).'''
    if not a or all(x == 0 for x in a):
        return 0
    if any(x < 0 for x in a):
        negWidth = max(((-x - 1).bit_length() for x in a if x < 0), default=0)
        posWidth = max((x.bit_length() for x in a if x > 0), default=0)
        return max(negWidth, posWidth) + 1
    return max(x.bit_length() for x in a)


def vectorConst(value: int, batchSize: int) -> list[int]:
    '''Constant-filled vector. Mirrors IntType(value, value, 0) initialization for a value-batch.'''
    if batchSize < 0:
        raise ValueError(f'vectorConst: batchSize must be non-negative, got {batchSize}')
    return [value] * batchSize


def formatNafExpr(naf: list[tuple[int, int]]) -> str:
    '''Format a NAF list [(sign, exponent), ...] as a string like "-2^91 + 2^43" or "1".'''
    if not naf:
        return '0'
    parts = []
    for i, (sign, k) in enumerate(naf):
        magnitude = '1' if k == 0 else f'2^{k}'
        if i == 0:
            parts.append(('-' if sign < 0 else '') + magnitude)
        else:
            parts.append((' + ' if sign > 0 else ' - ') + magnitude)
    return ''.join(parts)


def parseNafExpr(s: str) -> list[tuple[int, int]]:
    '''Parse a NAF expression string like "-2^91 + 2^43" or "2^39" or "1" into a list of (sign, exponent) tuples.'''
    compact = s.replace(' ', '')
    if not compact:
        raise ValueError(f'empty NAF expression')
    out: list[tuple[int, int]] = []
    for term in re.findall(r'[+-]?[^+-]+', compact):
        if not term:
            continue
        if term[0] == '-':
            sign, body = -1, term[1:]
        elif term[0] == '+':
            sign, body = 1, term[1:]
        else:
            sign, body = 1, term
        if '^' in body:
            base, exp = body.split('^')
            if base != '2':
                raise ValueError(f'unsupported base in NAF term {term!r}, only base-2 is supported')
            k = int(exp)
        else:
            v = int(body)
            if v <= 0 or (v & (v - 1)) != 0:
                raise ValueError(f'NAF term {term!r} must be a positive power of 2')
            k = v.bit_length() - 1
        out.append((sign, k))
    return out


# ---------------------------------------------------------------------------
# Spreadsheet I/O
#
# Every table this project keeps in xlsx is laid out the way `twiddles.xlsx` is:
# a header row, one data row per index, one column per series, and cells holding
# either a plain integer or a NAF expression string like "-2^91 + 2^43". A blank
# marks the end of the data. Writing replaces one named sheet and preserves the
# rest of the workbook.
#
# Those are the layout rules, and they live here so they are written once.
# Semantic wrappers stay with their consumers — `ntt.twiddles` for twiddle grids,
# `FullyPipelinedNTT.saveBoundsToXlsx` for bound tables, `scripts/build_bank.py`
# for constant columns — because which sheet means what is domain knowledge, not
# layout.
#
# openpyxl is imported inside the functions, so importing `utils` costs nothing
# for the callers that never touch a spreadsheet.
# ---------------------------------------------------------------------------


def _openpyxlModule():
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError(
            'openpyxl is required for xlsx I/O; install it with '
            '`pip install openpyxl` (it ships in the ntt-sage conda env)'
        ) from e
    return openpyxl


def _worksheet(path: str, sheetName: str):
    '''Open `path` read-only and return the named worksheet, or raise with the
    list of sheets that do exist — the error you actually want when a sheet name
    is wrong.'''
    openpyxl = _openpyxlModule()
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheetName not in wb.sheetnames:
        raise ValueError(f'sheet {sheetName!r} not found in {path!r}; available sheets: {wb.sheetnames}')
    return wb[sheetName]


def _headerCells(ws, headerRow: int, prefix: str | None) -> list:
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=headerRow, column=c).value
        if v is None:
            break
        if prefix is not None and not (isinstance(v, str)
                                       and v.strip().lower().startswith(prefix.lower())):
            break
        headers.append(v)
    return headers


def _dataRows(ws, headerRows: int, columns: int | None, firstColumn: int,
              stopOnBlank: bool) -> list[list]:
    if columns is None:
        columns = 0
        for c in range(firstColumn, ws.max_column + 1):
            if ws.cell(row=1, column=c).value is None:
                break
            columns += 1
    rows: list[list] = []
    for r in range(headerRows + 1, ws.max_row + 1):
        row = [ws.cell(row=r, column=firstColumn + i).value for i in range(columns)]
        if stopOnBlank and row and row[0] is None:
            break
        rows.append(row)
    return rows


def loadXlsxHeaders(path: str, sheetName: str, headerRow: int = 1,
                    prefix: str | None = None) -> list:
    '''Header cells of `sheetName`, left to right.

    Stops at the first blank cell, or — when `prefix` is given — at the first
    cell that is not a string starting with `prefix` (case-insensitively). That
    is how a grid separates its data columns from a trailing free-text label such
    as the "GS BUTTERFLIES!!!" that sits one column past the last "Layer N".
    '''
    return _headerCells(_worksheet(path, sheetName), headerRow, prefix)


def loadXlsxSheet(path: str, sheetName: str, headerRows: int = 1,
                  columns: int | None = None, firstColumn: int = 1,
                  stopOnBlank: bool = True) -> list[list]:
    '''Data rows of `sheetName` as raw cell values, one list per row.

    `headerRows` rows are skipped first — pass 0 for a sheet whose data starts at
    row 1. `columns` defaults to the number of contiguous non-blank cells in the
    sheet's first row. With `stopOnBlank` (the default) reading stops at the
    first row whose leading cell is empty, which is how these sheets mark their
    end; pass False to read through to `max_row`, blanks included.
    '''
    return _dataRows(_worksheet(path, sheetName), headerRows, columns,
                     firstColumn, stopOnBlank)


def loadXlsxColumn(path: str, sheetName: str, column: int = 1,
                   headerRows: int = 1, stopOnBlank: bool = True) -> list:
    '''One column of raw cell values — the flat-list case of `loadXlsxSheet`.'''
    return [row[0] for row in loadXlsxSheet(path, sheetName, headerRows=headerRows,
                                            columns=1, firstColumn=column,
                                            stopOnBlank=stopOnBlank)]


def saveXlsxSheet(path: str, sheetName: str, rows, headers=None,
                  position: int = 0) -> None:
    '''Write `rows` (a list of row lists) into `sheetName`.

    An existing sheet of that name is replaced and every other sheet in the
    workbook is preserved; the file is created when absent. `headers`, when
    given, is written verbatim across row 1 and the data starts at row 2.
    '''
    openpyxl = _openpyxlModule()
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        if sheetName in wb.sheetnames:
            del wb[sheetName]
        ws = wb.create_sheet(sheetName, position)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheetName

    firstDataRow = 1
    if headers is not None:
        for i, header in enumerate(headers):
            ws.cell(row=1, column=i + 1, value=header)
        firstDataRow = 2
    for r, row in enumerate(rows):
        for c, value in enumerate(row):
            ws.cell(row=firstDataRow + r, column=c + 1, value=value)
    wb.save(path)


def nafFromCell(value, sort: bool = True) -> list[tuple[int, int]]:
    '''NAF list from one spreadsheet cell: a plain integer, or a NAF expression
    string. Sorted ascending by exponent unless `sort=False` — ascending is the
    canonical order `calculateNttTwiddles` produces, so parsed cells and computed
    twiddles compare equal.'''
    if isinstance(value, bool):
        raise ValueError(f'expected an int or a NAF expression string, got {value!r}')
    if isinstance(value, int):
        naf = nafTerms(value)
    elif isinstance(value, str):
        naf = parseNafExpr(value)
    else:
        raise ValueError(f'expected an int or a NAF expression string, got {value!r}')
    if sort:
        naf.sort(key=lambda t: t[1])
    return naf


def nafToCell(value) -> str:
    '''Cell text for an integer or an existing NAF list, as `formatNafExpr`
    writes it — the form `nafFromCell` reads back.'''
    if isinstance(value, bool):
        raise TypeError(f'expected an int or a NAF list, got {value!r}')
    if isinstance(value, int):
        return formatNafExpr(nafTerms(value))
    if isinstance(value, list):
        return formatNafExpr(value)
    raise TypeError(f'expected an int or a NAF list, got {value!r}')


def loadXlsxNafGrid(path: str, sheetName: str, headerPrefix: str | None = None,
                    transpose: bool = False) -> list[list[list[tuple[int, int]]]]:
    '''A whole sheet of NAF cells, sized by its header row.

    Returns rows-by-columns as stored, or columns-by-rows with `transpose=True` —
    which is what a per-stage table wants, since those are written one column per
    stage and consumed one list per stage.
    '''
    ws = _worksheet(path, sheetName)
    headers = _headerCells(ws, 1, headerPrefix)
    if not headers:
        raise ValueError(
            f'sheet {sheetName!r} in {path!r} has no header cells'
            + (f' matching prefix {headerPrefix!r}' if headerPrefix else '')
            + ' to size the grid'
        )
    rows = _dataRows(ws, 1, len(headers), 1, True)

    # These sheets carry an optional free-text label one column past the last
    # data column ("GS BUTTERFLIES!!!"). A `headerPrefix` stops before it, but
    # without one the header row is wider than the data, so drop any trailing
    # column that is blank in every data row.
    width = len(headers)
    if rows:
        while width > 0 and all(row[width - 1] is None for row in rows):
            width -= 1
        if width == 0:
            raise ValueError(f'sheet {sheetName!r} in {path!r} has headers but no data cells')

    grid: list[list[list[tuple[int, int]]]] = []
    for r, row in enumerate(rows):
        parsed = []
        for c, value in enumerate(row[:width]):
            try:
                parsed.append(nafFromCell(value))
            except ValueError as e:
                raise ValueError(
                    f'sheet {sheetName!r} row {r + 2}, column {c + 1}: {e}'
                ) from e
        grid.append(parsed)

    if not transpose:
        return grid
    if not grid:
        return [[] for _ in headers]
    return [list(column) for column in zip(*grid)]


def saveXlsxNafGrid(grid, path: str, sheetName: str, headers=None,
                    label: str | None = None, transpose: bool = False) -> None:
    '''Inverse of `loadXlsxNafGrid`. `grid` holds ints or NAF lists; `transpose`
    means it is given columns-by-rows and must be written back rows-by-columns.
    `label` is the optional free-text cell one column past the last header.'''
    table = [list(row) for row in (zip(*grid) if transpose else grid)]
    if headers is not None:
        headers = list(headers)
        if label is not None:
            headers = headers + [label]
    rows = [[nafToCell(value) for value in row] for row in table]
    saveXlsxSheet(path, sheetName, rows, headers=headers)
