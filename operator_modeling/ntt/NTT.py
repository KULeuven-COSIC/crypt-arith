from __future__ import annotations
import os
import random
from math import log2
from sage.all import GF
from .Butterfly import Butterfly
from .ButterflyScheme import ButterflyScheme
from ..core.utils import nafTerms, nafTermsModulusLift, bitReverse, formatNafExpr, parseNafExpr
from ..core.IntType import IntType

# versal_arith path was added to sys.path by the .ButterflyScheme import above
# (see ButterflyScheme.py for the on-demand append). Both the
# pipeline spec dataclass and the rtl_gen ntt generator are imported through it.
from ntt_spec import NTTOperatorSpec, InterStageWire


def calculateNttTwiddles(modulus: int,
                         n: int,
                         butterflyType: str,
                         primitiveRoot: int | None = None,
                         negacyclic: bool = False,
                         useModulusLiftingNaf: bool = False,
                         maxPower: int = 95,
                         maxMultipleOfModulus: int = 2**32,
                         maxSearchDepth: int = 3,
                         beamWidth: int = 200,
                         maxNumberOfTerms: int = 96) -> list[list[int | list[tuple[int, int]]]]:
    '''Use sage to calculate the twiddle factors of the target NTT. Returns list[list[int | list[tuple[int, int]]]] of shape log2(n) x n/2: outer list indexes pipeline stages, inner list indexes butterflies within a stage in physical top-to-bottom order. If useModulusLiftingNaf is True, each twiddle is replaced by its NAF lift (a list of (sign, exponent) tuples). CT assumes bit-reversed input -> natural output (stride doubles each stage). GS assumes natural input -> bit-reversed output (stride halves each stage). Cyclic uses w = F.zeta(n) with linear exponent progression j * stepExp. Negacyclic uses psi = F.zeta(2n) with odd exponent progression (2j+1) * stepExp — only valid for forward NTT in CT direction (the GS butterfly equation does not admit a clean pre-twist absorption for forward NWC).'''
    if n < 2 or (n & (n - 1)) != 0:
        raise ValueError(f'n must be a power of 2 and at least 2, got {n}')
    if butterflyType not in ('CT', 'GS'):
        raise ValueError(f"butterflyType must be 'CT' or 'GS', got {butterflyType!r}")
    if negacyclic and butterflyType != 'CT':
        raise ValueError(
            f"negacyclic forward NTT requires butterflyType == 'CT'; "
            f"got {butterflyType!r}. The GS butterfly equation puts the twiddle "
            f"on the subtract branch's bOut, which prevents clean pre-twist absorption "
            f"into per-butterfly twiddles. Use CT for forward NWC and GS for inverse NWC."
        )

    F = GF(modulus)
    if primitiveRoot is None:
        base = F.zeta(2 * n) if negacyclic else F.zeta(n)
    else:
        base = primitiveRoot
    L = int(log2(n))

    twiddles: list[list[int | list[tuple[int, int]]]] = []
    for s in range(L):
        if butterflyType == 'CT':
            stride = 1 << s              # butterflies per group; doubles each stage
            groupSize = stride << 1      # 2^(s+1)
        else:                            # 'GS'
            groupSize = 1 << (L - s)     # halves each stage
            stride = groupSize >> 1
        repeats = n // groupSize         # group pattern repetition count to fill n/2 butterflies
        stepExp = n // groupSize         # exponent step (= n/groupSize)

        if negacyclic:
            # NWC merged twiddle in physical butterfly order: ψ^((2j+1) * stepExp).
            # Derived from path-product analysis: cyclic ω-exponent is j*stepExp; in ψ
            # that's 2j*stepExp; absorbing the per-stage pre-twist factor ψ^(2^(L-1-s)) =
            # ψ^stepExp adds +stepExp, giving (2j+1)*stepExp. Always odd, ranges over
            # the n distinct primitive 2n-th roots that aren't n-th roots.
            baseExps = [(2 * j + 1) * stepExp for j in range(stride)]
        else:
            baseExps = [j * stepExp for j in range(stride)]
        stageExps = baseExps * repeats

        stageTwiddles: list[int | list[tuple[int, int]]] = [int(base ** e) for e in stageExps]
        if useModulusLiftingNaf:
            stageTwiddles = [
                nafTermsModulusLift(t, modulus, maxPower, maxMultipleOfModulus,
                                    maxSearchDepth, beamWidth, maxNumberOfTerms)[1]
                for t in stageTwiddles
            ]
        twiddles.append(stageTwiddles)

    return twiddles


def calculateInttTwiddles(modulus: int,
                          n: int,
                          butterflyType: str,
                          primitiveRoot: int | None = None,
                          negacyclic: bool = False,
                          useModulusLiftingNaf: bool = False,
                          maxPower: int = 95,
                          maxMultipleOfModulus: int = 2**32,
                          maxSearchDepth: int = 3,
                          beamWidth: int = 200,
                          maxNumberOfTerms: int = 96) -> list[list[int | list[tuple[int, int]]]]:
    '''Inverse-NTT twiddle factors. Same shape and conventions as calculateNttTwiddles, but uses w^(-1) (cyclic) or psi^(-1) (negacyclic) as the base so the resulting pipeline computes the inverse transform — modulo the 1/n scaling, which is intentionally omitted; apply externally if needed. For negacyclic, only GS is valid (forward NWC uses CT, inverse NWC uses GS — the standard NWC pairing).'''
    if negacyclic and butterflyType != 'GS':
        raise ValueError(
            f"negacyclic inverse NTT requires butterflyType == 'GS'; "
            f"got {butterflyType!r}. Standard NWC pairing: forward = CT, inverse = GS. "
            f"The CT butterfly equation does not admit a clean post-twist absorption for inverse NWC."
        )
    F = GF(modulus)
    forwardBase = F(primitiveRoot) if primitiveRoot is not None else (F.zeta(2 * n) if negacyclic else F.zeta(n))
    inverseBase = int(forwardBase ** (-1))
    # For negacyclic, calculateNttTwiddles guards against GS — temporarily route through
    # cyclic semantics so we can build the inverse twiddles using the same (2j+1)*stepExp
    # progression but in GS direction. This is symmetric to the CT-forward case.
    if negacyclic:
        # Inverse NWC GS twiddle in physical order: ψ^(-(2j+1) * stepExp) where stepExp uses
        # GS group structure. Build directly here rather than calling calculateNttTwiddles.
        if n < 2 or (n & (n - 1)) != 0:
            raise ValueError(f'n must be a power of 2 and at least 2, got {n}')
        L = int(log2(n))
        psi = F(forwardBase)
        twiddles: list[list[int | list[tuple[int, int]]]] = []
        for s in range(L):
            groupSize = 1 << (L - s)
            stride = groupSize >> 1
            repeats = n // groupSize
            stepExp = n // groupSize
            baseExps = [(2 * j + 1) * stepExp for j in range(stride)]
            stageExps = baseExps * repeats
            stageTwiddles: list[int | list[tuple[int, int]]] = [int(psi ** (-e)) for e in stageExps]
            if useModulusLiftingNaf:
                stageTwiddles = [
                    nafTermsModulusLift(t, modulus, maxPower, maxMultipleOfModulus,
                                        maxSearchDepth, beamWidth, maxNumberOfTerms)[1]
                    for t in stageTwiddles
                ]
            twiddles.append(stageTwiddles)
        return twiddles
    return calculateNttTwiddles(modulus=modulus, n=n, butterflyType=butterflyType,
                                primitiveRoot=inverseBase, negacyclic=False,
                                useModulusLiftingNaf=useModulusLiftingNaf,
                                maxPower=maxPower, maxMultipleOfModulus=maxMultipleOfModulus,
                                maxSearchDepth=maxSearchDepth, beamWidth=beamWidth,
                                maxNumberOfTerms=maxNumberOfTerms)


def referenceNtt(x: list[int], modulus: int,
                 primitiveRoot: int | None = None,
                 negacyclic: bool = False) -> list[int]:
    '''Slow O(n^2) reference NTT for verification. Computes Y[k] for k in [0, n):
        cyclic:     Y[k] = sum_i x[i] * w^(i*k)         mod modulus
        negacyclic: Y[k] = sum_i x[i] * psi^(i*(2k+1))  mod modulus
    where w = F.zeta(n) and psi = F.zeta(2n) by default (matching calculateNttTwiddles).
    Output is a Python list[int] in [0, modulus), in natural order Y[0], Y[1], ..., Y[n-1].
    To compare against a pipeline whose output is in bit-reversed memory order
    (FullyPipelinedNTT with butterflyType='GS'), apply bitReverse to one side.'''
    if not isinstance(x, list) or not all(isinstance(v, int) for v in x):
        raise TypeError('x must be a list[int]')
    n = len(x)
    if n < 1:
        raise ValueError('x must be non-empty')
    F = GF(modulus)
    base = F(primitiveRoot) if primitiveRoot is not None else (F.zeta(2 * n) if negacyclic else F.zeta(n))
    if negacyclic:
        return [int(sum(F(x[i]) * base ** (i * (2 * k + 1)) for i in range(n))) for k in range(n)]
    return [int(sum(F(x[i]) * base ** (i * k) for i in range(n))) for k in range(n)]


def referenceIntt(y: list[int], modulus: int,
                  primitiveRoot: int | None = None,
                  negacyclic: bool = False,
                  divideByN: bool = True) -> list[int]:
    '''Slow O(n^2) reference inverse NTT for verification. Inverts referenceNtt by raising to w^(-1) (cyclic) or psi^(-1) (negacyclic). With divideByN=True (default), returns the canonical mathematical inverse so referenceIntt(referenceNtt(x)) == x. With divideByN=False, omits the 1/n factor so the result equals n*x mod modulus — useful for verifying FullyPipelinedINTT, which intentionally drops the 1/n scaling. Output is in natural order x[0], x[1], ..., x[n-1].'''
    if not isinstance(y, list) or not all(isinstance(v, int) for v in y):
        raise TypeError('y must be a list[int]')
    n = len(y)
    if n < 1:
        raise ValueError('y must be non-empty')
    F = GF(modulus)
    fwdBase = F(primitiveRoot) if primitiveRoot is not None else (F.zeta(2 * n) if negacyclic else F.zeta(n))
    invBase = fwdBase ** (-1)
    if negacyclic:
        x = [sum(F(y[k]) * invBase ** (i * (2 * k + 1)) for k in range(n)) for i in range(n)]
    else:
        x = [sum(F(y[k]) * invBase ** (i * k) for k in range(n)) for i in range(n)]
    if divideByN:
        nInv = F(n) ** (-1)
        x = [v * nInv for v in x]
    return [int(v) for v in x]


def loadTwiddlesFromXlsx(path: str, sheetName: str = 'NTT_TWIDDLES') -> list[list[list[tuple[int, int]]]]:
    '''Load twiddles from an xlsx file matching the project's layout. Row 1 holds "Layer 1".."Layer L" headers (an optional trailing label cell like "GS BUTTERFLIES!!!" is ignored). Rows 2..n/2+1 hold per-butterfly twiddles in physical top-to-bottom order, columns are stages. Each cell is either a plain integer or a NAF expression like "-2^91 + 2^43". Returns list[list[list[tuple[int, int]]]] of shape L x (n/2): every twiddle is materialized as a NAF list (matching calculateNttTwiddles output with useModulusLiftingNaf=True). Integer cells are converted via nafTerms.'''
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheetName not in wb.sheetnames:
        raise ValueError(f'sheet {sheetName!r} not found in {path!r}; available sheets: {wb.sheetnames}')
    ws = wb[sheetName]

    L = 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip().lower().startswith('layer'):
            L = c
        else:
            break
    if L == 0:
        raise ValueError(f'no "Layer N" headers found in row 1 of sheet {sheetName!r}')

    halfN = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is None:
            break
        halfN = r - 1

    twiddles: list[list[list[tuple[int, int]]]] = [[[] for _ in range(halfN)] for _ in range(L)]
    for r in range(halfN):
        for layerIdx in range(L):
            v = ws.cell(row=r + 2, column=layerIdx + 1).value
            if isinstance(v, int):
                naf = nafTerms(v)
            elif isinstance(v, str):
                naf = parseNafExpr(v)
            else:
                raise ValueError(f'unexpected cell type at row {r + 2}, column {layerIdx + 1} in sheet {sheetName!r}: {v!r}')
            # canonicalize term order to match calculateNttTwiddles (ascending exponent)
            naf.sort(key=lambda t: t[1])
            twiddles[layerIdx][r] = naf
    return twiddles


def saveTwiddlesToXlsx(twiddles: list[list[int | list[tuple[int, int]]]], path: str, butterflyType: str, sheetName: str = 'NTT_TWIDDLES') -> None:
    '''Save twiddles in the calculated format (output of calculateNttTwiddles) to an xlsx file in the project's layout. Row 1 is "Layer 1".."Layer L" plus a trailing "<TYPE> BUTTERFLIES!!!" label. Rows 2..n/2+1 hold per-butterfly twiddles, written as NAF expression strings (e.g. "-2^91 + 2^43" or "1") so that loadTwiddlesFromXlsx round-trips byte-for-byte. Integer inputs are converted to NAF via nafTerms before writing. If the target file exists the named sheet is replaced and other sheets are preserved; otherwise a new workbook is created.'''
    import openpyxl
    L = len(twiddles)
    if L == 0:
        raise ValueError('twiddles must have at least one stage')
    halfN = len(twiddles[0])
    if any(len(layer) != halfN for layer in twiddles):
        raise ValueError('all stages of twiddles must have the same length')

    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        if sheetName in wb.sheetnames:
            del wb[sheetName]
        ws = wb.create_sheet(sheetName, 0)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheetName

    for layerIdx in range(L):
        ws.cell(row=1, column=layerIdx + 1, value=f'Layer {layerIdx + 1}')
    ws.cell(row=1, column=L + 1, value=f'{butterflyType} BUTTERFLIES!!!')

    for r in range(halfN):
        for layerIdx in range(L):
            v = twiddles[layerIdx][r]
            if isinstance(v, int):
                naf = nafTerms(v)
            elif isinstance(v, list):
                naf = v
            else:
                raise TypeError(f'unexpected twiddle value at stage {layerIdx}, butterfly {r}: {v!r}')
            ws.cell(row=r + 2, column=layerIdx + 1, value=formatNafExpr(naf))

    wb.save(path)


def butterflyToMems(p: int, stride: int) -> tuple[int, int]:
    '''For a butterfly at physical position p in a stage with butterfly stride `stride`, return (memA, memB): the in-place memory positions its two inputs/outputs occupy.'''
    low = p & (stride - 1)
    high = p >> (stride.bit_length() - 1)
    mA = high * (stride << 1) + low
    return mA, mA + stride


def memToButterfly(m: int, stride: int) -> tuple[int, str]:
    '''For memory position m at the boundary after a stage whose butterflies have stride `stride`, return (butterflyIndex, port): which butterfly produced this position and on which output port.'''
    bitPos = stride.bit_length() - 1
    low = m & (stride - 1)
    bit = (m >> bitPos) & 1
    high = m >> (bitPos + 1)
    p = high * stride + low
    return p, 'A' if bit == 0 else 'B'


class FullyPipelinedNTT():
    def __init__(self, name: str, n: int, q: int, butterflyType: str, twiddles: list[list[int | list[tuple[int, int]]]], negacyclic: bool = False):
        self.name = name
        if n <= 0:
            raise ValueError(f"dimension n must be a positive number, got {n} instead")
        elif n & (n - 1) != 0:
            raise ValueError(f"dimension n must be a power of 2, got {n} instead")
        else:
            self.n = n
        if q <= 0:
            raise ValueError(f"q must be a positive number, got {q} instead")
        else:
            self.q = q
        if not isinstance(butterflyType, str):
            raise TypeError(f"butterflyType must be str, got {type(butterflyType)} instead")
        elif butterflyType == 'CT' or butterflyType == 'GS':
            self.butterflyType = butterflyType
        else:
            raise ValueError(f"butterflyType must be 'CT' or 'GS', got {butterflyType} instead")
        if not isinstance(twiddles, list):
            raise TypeError(f"twiddles must be a list of layer list containing int or naf tuples, got {type(twiddles)} instead")
        self.twiddles = twiddles
        self.negacyclic = negacyclic

        L = int(log2(n))
        self.butterflies: list[list[Butterfly]] = []
        for layerIndex in range(L):
            if self.butterflyType == 'CT':
                strideThis = 1 << layerIndex                # doubles each stage
                stridePrev = strideThis >> 1
            else:                                            # 'GS'
                strideThis = 1 << (L - layerIndex - 1)      # halves each stage
                stridePrev = strideThis << 1
            layer: list[Butterfly] = []
            for butterflyIndex in range(n // 2):
                bfly = Butterfly(
                    name=f"{self.name}_layer{layerIndex}_butterfly{butterflyIndex}",
                    butterflyType=self.butterflyType,
                    twiddle=self.twiddles[layerIndex][butterflyIndex],
                )
                if layerIndex > 0:
                    mA, mB = butterflyToMems(butterflyIndex, strideThis)
                    pA, portA = memToButterfly(mA, stridePrev)
                    pB, portB = memToButterfly(mB, stridePrev)
                    bfly.connectInTo(
                        connectATo=(self.butterflies[layerIndex - 1][pA], portA),
                        connectBTo=(self.butterflies[layerIndex - 1][pB], portB),
                    )
                layer.append(bfly)
            self.butterflies.append(layer)

    def getInputs(self, inputs: list[IntType] | list[list[int]]) -> None:
        '''Drive the first-stage butterfly input ports from `inputs`. The list is indexed by NTT memory position, in the layout the chosen scheme expects: bit-reversed natural order for CT (e.g. n=8: [x[0], x[4], x[2], x[6], x[1], x[5], x[3], x[7]]) and natural order for GS (e.g. n=8: [x[0], x[1], ..., x[7]]). Each first-stage butterfly's two memory slots are looked up via butterflyToMems and pulled from `inputs`.'''
        if not isinstance(inputs, list):
            raise TypeError(f'inputs must be a list, got {type(inputs)}')
        if len(inputs) != self.n:
            raise ValueError(f'inputs must have length n={self.n}, got {len(inputs)}')

        # Validate input mode: bound-mode (list[IntType]) or value-batch mode (list[list[int]]).
        if all(isinstance(x, list) for x in inputs):
            batchSize = len(inputs[0])
            if any(len(v) != batchSize for v in inputs):
                raise ValueError('all test-vector batches must have the same length; got mixed lengths in inputs')
        elif not all(isinstance(x, IntType) for x in inputs):
            raise TypeError('inputs must be either list[IntType] or list[list[int]]')

        stride0 = 1 if self.butterflyType == 'CT' else self.n // 2
        for p in range(self.n // 2):
            mA, mB = butterflyToMems(p, stride0)
            self.butterflies[0][p].initializeInputs(inputA=inputs[mA], inputB=inputs[mB])

    def getInputsNatural(self, x: list[IntType] | list[list[int]]) -> None:
        '''Drive the first-stage inputs from a list given in NATURAL order x[0], x[1], ..., x[n-1]. Auto-permutes into the pipeline's memory layout: bit-reverses for CT (which expects bit-reversed memory at input) or passes through for GS (already natural at input).'''
        if not isinstance(x, list):
            raise TypeError(f'x must be a list, got {type(x)}')
        if len(x) != self.n:
            raise ValueError(f'x must have length n={self.n}, got {len(x)}')
        L = int(log2(self.n))
        if self.butterflyType == 'CT':
            permuted = [x[bitReverse(m, L)] for m in range(self.n)]
        else:
            permuted = list(x)
        self.getInputs(permuted)

    def getOutputsNatural(self) -> list[IntType] | list[list[int]]:
        '''Return the final-stage outputs in NATURAL order Y[0], Y[1], ..., Y[n-1] (or x[0], ..., x[n-1] for INTT). Auto-permutes from the pipeline's memory layout: pass-through for CT (output is already natural) or bit-reverses for GS (output is bit-reversed in memory).'''
        L = int(log2(self.n))
        raw = self.getOutputs()
        if self.butterflyType == 'CT':
            return raw
        return [raw[bitReverse(k, L)] for k in range(self.n)]

    def setScheme(self, schemes: list[list[ButterflyScheme]]) -> None:
        '''Attach a ButterflyScheme to every butterfly. `schemes` is shaped log2(n) x n/2, mirroring the butterfly grid: schemes[layerIndex][butterflyIndex] is assigned to self.butterflies[layerIndex][butterflyIndex].scheme.'''
        if not isinstance(schemes, list):
            raise TypeError(f'schemes must be a list of layer lists, got {type(schemes)}')
        L = len(self.butterflies)
        if len(schemes) != L:
            raise ValueError(f'schemes must have {L} layers (log2(n)), got {len(schemes)}')
        halfN = self.n // 2
        for layerIndex, layer in enumerate(schemes):
            if not isinstance(layer, list):
                raise TypeError(f'schemes[{layerIndex}] must be a list, got {type(layer)}')
            if len(layer) != halfN:
                raise ValueError(f'schemes[{layerIndex}] must have n/2={halfN} entries, got {len(layer)}')
            for butterflyIndex, scheme in enumerate(layer):
                if not isinstance(scheme, ButterflyScheme):
                    raise TypeError(f'schemes[{layerIndex}][{butterflyIndex}] must be a ButterflyScheme, got {type(scheme)}')
                self.butterflies[layerIndex][butterflyIndex].scheme = scheme

    def compute(self) -> None:
        '''Walk the pipeline stage by stage, calling each butterfly's compute(). Each butterfly pulls input bounds from its input ports (driven by getInputs() for stage 0, or by upstream butterflies' compute() pushes for later stages), runs its scheme, and pushes its output bounds onto downstream input ports. By the time stage s runs, stage s-1 has already populated all stage-s input ports.'''
        for layer in self.butterflies:
            for bfly in layer:
                bfly.compute()

    def getOutputs(self) -> list[IntType] | list[list[int]]:
        '''Return the final-layer outputs in NTT memory-position order. Mirror of getInputs: outputs[m] is the value at memory slot m. Auto-dispatches per-port: if outputPort.bound is set it's used; otherwise outputPort.testVector. Result is None for any slot whose owning butterfly has not had compute() run yet. Return type is list[IntType] in bound mode and list[list[int]] in value-batch mode (mixed populations follow per-port).'''
        L = len(self.butterflies)
        strideLast = (self.n // 2) if self.butterflyType == 'CT' else 1
        outputs: list = [None] * self.n
        for m in range(self.n):
            p, port = memToButterfly(m, strideLast)
            bfly = self.butterflies[L - 1][p]
            outPort = bfly.outputPortA if port == 'A' else bfly.outputPortB
            outputs[m] = outPort.bound if outPort.bound is not None else outPort.testVector
        return outputs

    def showBounds(self) -> None:
        '''Print a per-stage summary of output bounds: max bitWidth across the stage's butterflies, plus a representative IntType (the widest one). Quick at-a-glance view of how bounds evolve through the pipeline. Run after compute().'''
        print(f'=== Bounds for {self.name} (n={self.n}, {self.butterflyType}) ===')
        for s, layer in enumerate(self.butterflies):
            widest: IntType | None = None
            maxWidth = -1
            for bfly in layer:
                for bound in (bfly.outputPortA.bound, bfly.outputPortB.bound):
                    if bound is None:
                        continue
                    if bound.bitWidth > maxWidth:
                        maxWidth = bound.bitWidth
                        widest = bound
            if widest is None:
                print(f'Layer {s}: <no bounds — has compute() been run?>')
            else:
                print(f'Layer {s}: max bitWidth = {maxWidth}, e.g. {widest}')

    def saveBoundsToXlsx(self, path: str = 'NTT_bounds.xlsx', sheetName: str | None = None) -> None:
        '''Record the full per-butterfly per-stage output bound table to xlsx. Layout: row 1 holds "Layer 1".."Layer L" headers plus a trailing "<TYPE> BOUNDS" label; rows 2..n+1 hold per-port bounds in physical order (butterfly p port A on row 2+2p, port B on row 3+2p). Each cell contains str(bound) using IntType.__str__. Sheet name defaults to self.name so multiple NTT instances can record to the same workbook without colliding. If the file exists, the named sheet is replaced and other sheets are preserved; otherwise a new workbook is created.'''
        import openpyxl
        if sheetName is None:
            sheetName = self.name
        L = len(self.butterflies)

        if os.path.exists(path):
            wb = openpyxl.load_workbook(path)
            if sheetName in wb.sheetnames:
                del wb[sheetName]
            ws = wb.create_sheet(sheetName, 0)
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheetName

        for layerIdx in range(L):
            ws.cell(row=1, column=layerIdx + 1, value=f'Layer {layerIdx + 1}')
        ws.cell(row=1, column=L + 1, value=f'{self.butterflyType} BOUNDS')

        for p in range(self.n // 2):
            for layerIdx in range(L):
                bfly = self.butterflies[layerIdx][p]
                aBound = bfly.outputPortA.bound
                bBound = bfly.outputPortB.bound
                ws.cell(row=2 + 2 * p, column=layerIdx + 1,
                        value=str(aBound) if aBound is not None else '')
                ws.cell(row=3 + 2 * p, column=layerIdx + 1,
                        value=str(bBound) if bBound is not None else '')

        wb.save(path)

    def getOperatorInterface(self, name: str) -> NTTOperatorSpec:
        '''Build an NTTOperatorSpec describing the entire pipeline at RTL granularity.

        Walks the populated grid (compute() must already have been run with input
        bounds loaded so every input/output port has a bound), invokes each
        butterfly's scheme.getOperatorInterface to get its ButterflyOperatorSpec,
        and precomputes the wiring tables that map natural-order x[i] inputs to
        first-stage butterfly ports (inputWiring), inter-layer connections
        (interStageWiring via butterflyToMems + memToButterfly), and final-stage
        butterfly ports back to natural-order y[i] outputs (outputWiring).

        Per-butterfly module names are namespaced by the NTT's top name as
        <name>_btf_L<s>_p<p>, so two NTT instances with the same n /
        butterflyType but different `name` produce non-colliding butterfly
        modules — required when integrating multiple NTTs into one Vivado
        project. The short `btf` abbreviation (vs full `Butterfly`) keeps
        the resulting module + instance identifiers short enough to dodge a
        Vivado xelab false-positive multi-driver bug that fires on full
        449-butterfly elaboration when individual identifiers grow past
        ~50 characters. Standalone build_butterfly.py keeps its own
        Butterfly_n<N>_<TYPE>_L<s>_p<p> shape — single-module compiles do
        not trigger that pathology.

        Consumed by versal_arith.rtl_gen.ntt.NTT_RTL_gen, which composes the
        per-butterfly RTL into a full pipeline wrapper.
        '''
        n = self.n
        L = len(self.butterflies)
        if L != int(log2(n)):
            raise ValueError(f'butterfly grid layer count {L} does not match log2(n)={int(log2(n))}')
        for s in range(L):
            if len(self.butterflies[s]) != n // 2:
                raise ValueError(f'butterflies[{s}] has {len(self.butterflies[s])} entries, expected {n // 2}')

        # Per-natural-input widths come from each stage-0 butterfly port's bound. Filled
        # below after `inputWiring` is computed. Each stage-0 butterfly's two input ports
        # must have a bound set (by getInputsNatural + compute()).
        for p in range(n // 2):
            for portName, port in (('A', self.butterflies[0][p].inputPortA),
                                   ('B', self.butterflies[0][p].inputPortB)):
                if port.bound is None:
                    raise ValueError(f'butterflies[0][{p}].inputPort{portName}.bound is None — call compute() with bounds first')

        # Per-butterfly ButterflyOperatorSpec collection.
        butterflySpecs: list[list] = []
        for s in range(L):
            layerSpecs = []
            for p in range(n // 2):
                bfly = self.butterflies[s][p]
                if bfly.scheme is None:
                    raise ValueError(f'butterflies[{s}][{p}].scheme is None — call setScheme() before getOperatorInterface()')
                if bfly.inputPortA.bound is None or bfly.inputPortB.bound is None:
                    raise ValueError(f'butterflies[{s}][{p}].inputPort{{A,B}}.bound is None — call compute() with bounds first')
                # The scheme's aIn/bIn may have been left as testVectors by the last compute()
                # call (when both bounds and values were loaded). Reset to bounds so
                # getOperatorInterface inspects the right state.
                bfly.scheme.aIn = bfly.inputPortA.bound
                bfly.scheme.bIn = bfly.inputPortB.bound
                if bfly.twiddle is not None:
                    bfly.scheme.twiddle = bfly.twiddle
                elif bfly.twiddlePort is not None:
                    bfly.scheme.twiddle = bfly.twiddlePort.bound
                else:
                    raise ValueError(f'butterflies[{s}][{p}] has no twiddle assigned')
                butterflyName = f'{name}_btf_L{s}_p{p}'
                layerSpecs.append(bfly.scheme.getOperatorInterface(name=butterflyName))
            butterflySpecs.append(layerSpecs)

        # Stage-0 routing in natural-order space. Mirrors getInputsNatural's
        # permutation logic: CT bit-reverses the natural input before driving
        # memory slots, GS passes through. Each natural-order x[i]'s width
        # comes from the corresponding stage-0 butterfly port's bound, with no
        # uniformity assumption.
        stride0 = 1 if self.butterflyType == 'CT' else n // 2
        inputWiring: list[tuple[int, int]] = []
        inputBitWidthsNatural: list[int] = [0] * n
        inputIsSignedNatural: list[bool] = [False] * n
        for p in range(n // 2):
            mA, mB = butterflyToMems(p, stride0)
            if self.butterflyType == 'CT':
                natA, natB = bitReverse(mA, L), bitReverse(mB, L)
            else:
                natA, natB = mA, mB
            inputWiring.append((natA, natB))
            aInBound = self.butterflies[0][p].inputPortA.bound
            bInBound = self.butterflies[0][p].inputPortB.bound
            inputBitWidthsNatural[natA] = aInBound.bitWidth
            inputIsSignedNatural[natA] = aInBound.isSigned
            inputBitWidthsNatural[natB] = bInBound.bitWidth
            inputIsSignedNatural[natB] = bInBound.isSigned

        # Final-stage routing in natural-order space. Mirrors getOutputsNatural:
        # CT outputs are already natural, GS outputs are bit-reversed.
        strideLast = (n // 2) if self.butterflyType == 'CT' else 1
        outputWiring: list[tuple[int, int]] = []
        outputBitWidthsNatural: list[int] = [0] * n
        outputIsSignedNatural: list[bool] = [False] * n
        for p in range(n // 2):
            mA, mB = butterflyToMems(p, strideLast)
            if self.butterflyType == 'CT':
                natA, natB = mA, mB
            else:
                natA, natB = bitReverse(mA, L), bitReverse(mB, L)
            outputWiring.append((natA, natB))
            aBound = self.butterflies[L - 1][p].outputPortA.bound
            bBound = self.butterflies[L - 1][p].outputPortB.bound
            if aBound is None or bBound is None:
                raise ValueError(f'butterflies[{L-1}][{p}].outputPort{{A,B}}.bound is None — call compute() with bounds first')
            outputBitWidthsNatural[natA] = aBound.bitWidth
            outputIsSignedNatural[natA] = aBound.isSigned
            outputBitWidthsNatural[natB] = bBound.bitWidth
            outputIsSignedNatural[natB] = bBound.isSigned

        # Inter-stage wiring for layers 1..L-1. Replicates the connection logic from
        # FullyPipelinedNTT.__init__ (see NTT.py:299-323): use this layer's stride
        # to find the butterfly's memory positions, then the previous layer's
        # stride to find which upstream butterfly+port produced each position.
        interStageWiring: list[list[tuple[InterStageWire, InterStageWire]]] = []
        for s in range(1, L):
            if self.butterflyType == 'CT':
                strideThis = 1 << s
                stridePrev = strideThis >> 1
            else:
                strideThis = 1 << (L - s - 1)
                stridePrev = strideThis << 1
            layerWiring: list[tuple[InterStageWire, InterStageWire]] = []
            for p in range(n // 2):
                mA, mB = butterflyToMems(p, strideThis)
                pA, portA = memToButterfly(mA, stridePrev)
                pB, portB = memToButterfly(mB, stridePrev)
                layerWiring.append((InterStageWire(src_p=pA, src_port=portA),
                                    InterStageWire(src_p=pB, src_port=portB)))
            interStageWiring.append(layerWiring)

        return NTTOperatorSpec(
            name=name,
            n=n,
            butterflyType=self.butterflyType,
            negacyclic=self.negacyclic,
            q=self.q,
            butterflySpecs=butterflySpecs,
            inputBitWidthsNatural=inputBitWidthsNatural,
            inputIsSignedNatural=inputIsSignedNatural,
            outputBitWidthsNatural=outputBitWidthsNatural,
            outputIsSignedNatural=outputIsSignedNatural,
            inputWiring=inputWiring,
            outputWiring=outputWiring,
            interStageWiring=interStageWiring,
        )

    def _extractGoldensNatural(self) -> tuple[list[list[int]], list[list[int]]]:
        '''Walk the populated grid and return (goldenXNatural, goldenYNatural),
        each shape (testSize x n). Inputs come from stage-0 input ports'
        testVector (set by getInputsNatural([batches])); outputs come from
        final-layer output ports' testVector (set by compute()). The
        natural ↔ memory permutation reuses the same mapping that
        getInputsNatural / getOutputsNatural use. Precondition: every relevant
        port has a testVector populated (call getInputsNatural([list[int]]*n)
        and compute() before invoking this).'''
        n = self.n
        L = len(self.butterflies)
        stride0 = 1 if self.butterflyType == 'CT' else n // 2
        strideLast = (n // 2) if self.butterflyType == 'CT' else 1

        inputByNatural: list[list[int] | None] = [None] * n
        for p in range(n // 2):
            mA, mB = butterflyToMems(p, stride0)
            if self.butterflyType == 'CT':
                natA, natB = bitReverse(mA, L), bitReverse(mB, L)
            else:
                natA, natB = mA, mB
            inputByNatural[natA] = self.butterflies[0][p].inputPortA.testVector
            inputByNatural[natB] = self.butterflies[0][p].inputPortB.testVector

        outputByNatural: list[list[int] | None] = [None] * n
        for p in range(n // 2):
            mA, mB = butterflyToMems(p, strideLast)
            if self.butterflyType == 'CT':
                natA, natB = mA, mB
            else:
                natA, natB = bitReverse(mA, L), bitReverse(mB, L)
            outputByNatural[natA] = self.butterflies[L - 1][p].outputPortA.testVector
            outputByNatural[natB] = self.butterflies[L - 1][p].outputPortB.testVector

        for i in range(n):
            if inputByNatural[i] is None:
                raise ValueError(
                    f'natural-input index {i}: stage-0 butterfly port has no testVector. '
                    f'Call getInputsNatural([list[int]]*n) + compute() before emitRtl.'
                )
            if outputByNatural[i] is None:
                raise ValueError(
                    f'natural-output index {i}: final-layer butterfly port has no testVector. '
                    f'Call compute() with values loaded before emitRtl.'
                )

        testSize = len(inputByNatural[0])
        for i in range(n):
            if len(inputByNatural[i]) != testSize:
                raise ValueError(
                    f'inconsistent batch length on natural-input index {i}: '
                    f'len={len(inputByNatural[i])}, expected {testSize}'
                )
            if len(outputByNatural[i]) != testSize:
                raise ValueError(
                    f'inconsistent batch length on natural-output index {i}: '
                    f'len={len(outputByNatural[i])}, expected {testSize}'
                )

        goldenXNatural = [[inputByNatural[i][b] for i in range(n)] for b in range(testSize)]
        goldenYNatural = [[outputByNatural[i][b] for i in range(n)] for b in range(testSize)]
        return goldenXNatural, goldenYNatural

    def emitRtl(self,
                topName: str,
                run_dir,
                pipeline_stages_per_layer = 1,
                gen_testbench: bool = True,
                visualization: bool = False,
                sanity_check_size: int = 8,
                backend: str = 'hw') -> dict:
        '''Emit RTL for the entire NTT/INTT pipeline. Precondition: setScheme
        was called, getInputsNatural([bounds]) + compute() populated the bound
        path. When gen_testbench=True, also requires getInputsNatural([batches])
        + compute() so every input/output port has a testVector — the goldens
        come from the populated instance, not internal sampling.

        `pipeline_stages_per_layer` is either an int (broadcast to every layer)
        or a list of length log2(n).

        `backend` selects the RTL flavor:
          - 'hw'  (default): optimized Versal compressor-tree generator
            (`rtl_gen.ntt.NTT_RTL_gen`). Per-butterfly `.sv` files, full
            `xdc_generated/` directory, bit-heap intermediate `.txt` files.
          - 'sim': behavioral simulation-only generator
            (`sim_rtl_gen.ntt.NTT_SimRTL_gen`). All butterfly modules
            concatenated into a single `<topName>_butterflies.sv`; no
            `xdc_generated/`, no bit-heap files. Testvectors are
            byte-identical to the hw backend.

        Files land in `<run_dir>/RTL_generated/`, `<run_dir>/xdc_generated/`
        (hw only), `<run_dir>/testvectors/` (when gen_testbench),
        `<run_dir>/manifest.json`.

        A local sanity check (off when sanity_check_size <= 0) round-trips the
        first `sanity_check_size` testvector lines to confirm the on-disk hex
        matches what propagateValue produces.'''
        import os as _os
        from pathlib import Path as _Path

        L = len(self.butterflies)
        if isinstance(pipeline_stages_per_layer, int):
            pipeline_stages_per_layer = [pipeline_stages_per_layer] * L
        if len(pipeline_stages_per_layer) != L:
            raise ValueError(
                f'pipeline_stages_per_layer must have length log2(n)={L}, '
                f'got {len(pipeline_stages_per_layer)}'
            )

        run_dir = _Path(run_dir)
        run_dir.mkdir(parents=True, exist_ok=True)

        spec = self.getOperatorInterface(name=topName)

        goldenXNatural = None
        goldenYNatural = None
        testSize = 0
        if gen_testbench:
            goldenXNatural, goldenYNatural = self._extractGoldensNatural()
            testSize = len(goldenXNatural)

        if backend == 'sim':
            from sim_rtl_gen.ntt import NTT_SimRTL_gen as _gen
        elif backend == 'hw':
            from rtl_gen.ntt import NTT_RTL_gen as _gen
        else:
            raise ValueError(f"backend must be 'hw' or 'sim', got {backend!r}")

        saved = _os.getcwd()
        _os.chdir(str(run_dir))
        try:
            manifest = _gen(
                spec=spec,
                pipeline_stages_per_layer=list(pipeline_stages_per_layer),
                gen_testbench=gen_testbench,
                test_size=testSize,
                visualization=visualization,
                golden_x_natural=goldenXNatural,
                golden_y_natural=goldenYNatural,
            )
        finally:
            _os.chdir(saved)

        if gen_testbench and sanity_check_size > 0:
            _sanityCheckNttTestvectors(self, run_dir, spec, sanity_check_size)

        return manifest


def _sanityCheckNttTestvectors(inst: 'FullyPipelinedNTT', run_dir, spec, sample_size: int = 8) -> None:
    '''Decode the first `sample_size` lines of x_in.txt and y_out.txt, drive
    the decoded x batch through the populated instance, and confirm each
    natural-output slot matches mod 2^slot_width. Catches twos-complement
    encoding bugs locally before any remote sim. Mirrors
    `scripts/build_ntt.py::sanity_check_ntt`.'''
    from pathlib import Path as _Path
    n = inst.n
    L = len(inst.butterflies)
    inWidths = list(spec.inputBitWidthsNatural)
    inSigned = list(spec.inputIsSignedNatural)
    outWidths = list(spec.outputBitWidthsNatural)
    tv = _Path(run_dir) / 'testvectors'

    def _readHex(path, nLines):
        with path.open() as f:
            return [int(line.strip(), 16) for _, line in zip(range(nLines), f) if line.strip()]

    xPacked = _readHex(tv / 'x_in.txt', sample_size)
    yPacked = _readHex(tv / 'y_out.txt', sample_size)
    nLines = min(len(xPacked), len(yPacked))
    if nLines == 0:
        raise RuntimeError('emitRtl sanity-check: no testvectors loaded from disk')

    def _decodePerSlotSigned(packed):
        out = []
        offset = 0
        for w, signed in zip(inWidths, inSigned):
            v = (packed >> offset) & ((1 << w) - 1)
            if signed and (v >> (w - 1)):
                v -= (1 << w)
            out.append(v)
            offset += w
        return out

    def _decodePerSlotUnsigned(packed):
        out = []
        offset = 0
        for w in outWidths:
            v = (packed >> offset) & ((1 << w) - 1)
            out.append(v)
            offset += w
        return out

    xDecoded = [_decodePerSlotSigned(p) for p in xPacked[:nLines]]
    yDecoded = [_decodePerSlotUnsigned(p) for p in yPacked[:nLines]]

    inputBoundsByNatural: list = [None] * n
    stride0 = 1 if inst.butterflyType == 'CT' else n // 2
    for p in range(n // 2):
        mA, mB = butterflyToMems(p, stride0)
        if inst.butterflyType == 'CT':
            natA, natB = bitReverse(mA, L), bitReverse(mB, L)
        else:
            natA, natB = mA, mB
        inputBoundsByNatural[natA] = inst.butterflies[0][p].inputPortA.bound
        inputBoundsByNatural[natB] = inst.butterflies[0][p].inputPortB.bound

    nInputs = [[xDecoded[b][i] for b in range(nLines)] for i in range(n)]
    inst.getInputsNatural(inputBoundsByNatural)
    inst.getInputsNatural(nInputs)
    inst.compute()

    strideLast = (n // 2) if inst.butterflyType == 'CT' else 1
    pipelineOutMemoryOrder: list = [None] * n
    for m in range(n):
        p, port = memToButterfly(m, strideLast)
        bfly = inst.butterflies[L - 1][p]
        outPort = bfly.outputPortA if port == 'A' else bfly.outputPortB
        pipelineOutMemoryOrder[m] = outPort.testVector
    if inst.butterflyType == 'CT':
        pipelineOutNatural = pipelineOutMemoryOrder
    else:
        pipelineOutNatural = [pipelineOutMemoryOrder[bitReverse(k, L)] for k in range(n)]

    mismatches = 0
    for b in range(nLines):
        for i in range(n):
            expected = pipelineOutNatural[i][b] & ((1 << outWidths[i]) - 1)
            actual = yDecoded[b][i]
            if expected != actual:
                mismatches += 1
                if mismatches <= 5:
                    print(f'[emitRtl] sanity-check mismatch batch={b} idx={i}: '
                          f'expected {expected:x}, got {actual:x}')
    if mismatches > 0:
        raise RuntimeError(
            f'emitRtl sanity-check FAILED: {mismatches} mismatches out of '
            f'{nLines * n} slots — twos-complement encoding bug?'
        )
    print(f'[emitRtl] sanity-check OK: {nLines} batches × {n} slots = {nLines * n} OK')


class FullyPipelinedINTT(FullyPipelinedNTT):
    '''Inverse fully-pipelined NTT. Identical to FullyPipelinedNTT in wiring, scheme attachment, compute pipeline, and output reporting; the inverse-ness lives entirely in the twiddles passed in (use calculateInttTwiddles to generate them with w^(-1) or psi^(-1) as the base). The 1/n scaling factor at the end of the inverse transform is intentionally omitted here; apply it externally if needed.'''
    pass


def _verifyImpl(instance: FullyPipelinedNTT, isInverse: bool, primitiveRoot: int | None,
                batchSize: int, seed: int | None,
                inputBound: 'IntType | list[IntType] | None',
                valueRange: tuple[int, int] | None, verbose: bool) -> bool:
    if instance.negacyclic:
        # NWC requires the standard pairing: forward = CT, inverse = GS.
        if not isInverse and instance.butterflyType != 'CT':
            raise ValueError(
                f"verifyNtt for negacyclic instance requires butterflyType == 'CT'; "
                f"got {instance.butterflyType!r}. The GS butterfly equation does not "
                f"compute a forward NWC NTT under cyclic-CT-style wiring."
            )
        if isInverse and instance.butterflyType != 'GS':
            raise ValueError(
                f"verifyIntt for negacyclic instance requires butterflyType == 'GS'; "
                f"got {instance.butterflyType!r}. The CT butterfly equation does not "
                f"compute an inverse NWC NTT under cyclic-CT-style wiring."
            )
    if seed is not None:
        random.seed(seed)
    n = instance.n
    q = instance.q
    negacyclic = instance.negacyclic

    # inputBound accepts three forms:
    #   - None         : default IntType.signed(66) broadcast to every natural index
    #   - IntType      : broadcast to every natural index (current behavior)
    #   - list[IntType]: per-natural-index list of length n; each x[i]'s value
    #                    range is derived from its own bound's [minValue, maxValue]
    #                    and `valueRange` is ignored (a single global range cannot
    #                    represent per-port mixed signed/unsigned widths)
    if inputBound is None:
        inputBound = IntType.signed(66)
    if isinstance(inputBound, list):
        if len(inputBound) != n:
            raise ValueError(
                f'inputBound list must have length n={n}, got {len(inputBound)}'
            )
        bounds_per_natural = inputBound
        natural = [[random.randint(b.minValue, b.maxValue) for _ in range(batchSize)]
                   for b in bounds_per_natural]
    else:
        if valueRange is None:
            valueRange = (-(2 ** 65), 2 ** 65 - 1)
        bounds_per_natural = [inputBound] * n
        natural = [[random.randint(valueRange[0], valueRange[1]) for _ in range(batchSize)]
                   for _ in range(n)]

    instance.getInputsNatural(bounds_per_natural)
    instance.getInputsNatural(natural)
    instance.compute()

    # Read testVector directly from final-layer output ports in memory order, then
    # permute to natural order. We can't use getOutputsNatural() here because
    # getOutputs() prefers bound when both are set — and we set both above.
    L = len(instance.butterflies)
    strideLast = (n // 2) if instance.butterflyType == 'CT' else 1
    pipelineOutMemoryOrder: list[list[int] | None] = [None] * n
    for m in range(n):
        p, port = memToButterfly(m, strideLast)
        bfly = instance.butterflies[L - 1][p]
        outPort = bfly.outputPortA if port == 'A' else bfly.outputPortB
        pipelineOutMemoryOrder[m] = outPort.testVector
    Llog = int(log2(n))
    if instance.butterflyType == 'CT':
        pipelineOutNatural = pipelineOutMemoryOrder
    else:
        pipelineOutNatural = [pipelineOutMemoryOrder[bitReverse(k, Llog)] for k in range(n)]

    modQFails = 0
    for b in range(batchSize):
        x = [natural[m][b] for m in range(n)]
        if isInverse:
            ref = referenceIntt(x, q, primitiveRoot=primitiveRoot, negacyclic=negacyclic, divideByN=False)
        else:
            ref = referenceNtt(x, q, primitiveRoot=primitiveRoot, negacyclic=negacyclic)
        for m in range(n):
            actual = pipelineOutNatural[m][b] % q
            if actual != ref[m]:
                modQFails += 1
                if verbose and modQFails <= 5:
                    print(f'  mod-q mismatch batch {b} pos {m}: actual={actual}, ref={ref[m]}')

    boundFails = 0
    portCount = 0
    for layer in instance.butterflies:
        for bfly in layer:
            for outPort in (bfly.outputPortA, bfly.outputPortB):
                bnd = outPort.bound
                vec = outPort.testVector
                if bnd is None or vec is None:
                    continue
                portCount += len(vec)
                for v in vec:
                    if v < bnd.minValue or v > bnd.maxValue:
                        boundFails += 1

    totalModQ = batchSize * n
    if verbose:
        label = 'verifyIntt' if isInverse else 'verifyNtt'
        print(f'{label} {instance.name}: '
              f'mod-q {totalModQ - modQFails}/{totalModQ}, '
              f'bound containment {portCount - boundFails}/{portCount}')
    return modQFails == 0 and boundFails == 0


def verifyNtt(instance: FullyPipelinedNTT,
              primitiveRoot: int | None = None,
              batchSize: int = 4,
              seed: int | None = None,
              inputBound: 'IntType | list[IntType] | None' = None,
              valueRange: tuple[int, int] | None = None,
              verbose: bool = True) -> bool:
    '''Generate `batchSize` random natural-order test vectors, run them through `instance` (loading both bounds and values so propagateValue uses the same hardware-register slicing as propagateBound), compute the reference via referenceNtt, and check (a) pipeline_output[m] mod q == referenceNtt(x)[m] for every m and every batch element, and (b) every actual test value at every output port lies inside that port's bound interval. Returns True iff both pass everywhere; False otherwise. With verbose=True (default) prints the first 5 mod-q failure lines and a summary. `primitiveRoot` should match what was used to build the twiddles; if None, defaults to F.zeta(n) (cyclic) or F.zeta(2n) (negacyclic), matching calculateNttTwiddles. `inputBound` accepts three forms: None (default IntType.signed(66) broadcast), a single IntType (broadcast to every natural index), or a list[IntType] of length n (per-natural-index bounds — values sampled within each bound's [minValue, maxValue], `valueRange` is ignored in this case). `valueRange` defaults to (-(2**65), 2**65 - 1) when broadcasting; override both consistently if changing either.'''
    return _verifyImpl(instance=instance, isInverse=False, primitiveRoot=primitiveRoot,
                       batchSize=batchSize, seed=seed, inputBound=inputBound,
                       valueRange=valueRange, verbose=verbose)


def verifyIntt(instance: FullyPipelinedINTT,
               primitiveRoot: int | None = None,
               batchSize: int = 4,
               seed: int | None = None,
               inputBound: 'IntType | list[IntType] | None' = None,
               valueRange: tuple[int, int] | None = None,
               verbose: bool = True) -> bool:
    '''Same as verifyNtt but compares against referenceIntt(..., divideByN=False), since FullyPipelinedINTT intentionally drops the 1/n scaling. Pass the FORWARD `primitiveRoot` (referenceIntt inverts it internally to get the inverse base). `inputBound` accepts the same three forms as verifyNtt: None, a single IntType, or a list[IntType] of length n.'''
    return _verifyImpl(instance=instance, isInverse=True, primitiveRoot=primitiveRoot,
                       batchSize=batchSize, seed=seed, inputBound=inputBound,
                       valueRange=valueRange, verbose=verbose)

