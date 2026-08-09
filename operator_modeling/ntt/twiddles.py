'''Twiddle factors: computing them, and moving them to and from a spreadsheet.

Split out of `NTT.py` for a concrete reason. That module did
`from sage.all import GF` unconditionally, so *any* import of it paid for Sage —
including `scripts/build_butterfly.py`, which only ever wanted
`loadTwiddlesFromXlsx`. The xlsx half needs openpyxl and no Sage; the
computation half needs Sage and no openpyxl. Separating them lets each caller
pay only for what it uses.
'''
from __future__ import annotations

from math import log2

from sage.all import GF

from ..core.utils import (bitReverse, formatNafExpr, nafTerms,
                          nafTermsModulusLift, parseNafExpr)

def calculateNttTwiddles(modulus: int,
                         n: int,
                         butterflyType: str,
                         primitiveRoot: int | None = None,
                         negacyclic: bool = False,
                         useModulusLiftingNaf: bool = False,
                         maxPower: int = 95,
                         maxMultipleOfModulus: int = 2**32,
                         maxSearchDepth: int = 3,
                         beamWidth: int = 200,
                         maxNumberOfTerms: int = 96) -> list[list[int | list[tuple[int, int]]]]:
    '''Use sage to calculate the twiddle factors of the target NTT. Returns list[list[int | list[tuple[int, int]]]] of shape log2(n) x n/2: outer list indexes pipeline stages, inner list indexes butterflies within a stage in physical top-to-bottom order. If useModulusLiftingNaf is True, each twiddle is replaced by its NAF lift (a list of (sign, exponent) tuples). CT assumes bit-reversed input -> natural output (stride doubles each stage). GS assumes natural input -> bit-reversed output (stride halves each stage). Cyclic uses w = F.zeta(n) with linear exponent progression j * stepExp. Negacyclic uses psi = F.zeta(2n) with odd exponent progression (2j+1) * stepExp — only valid for forward NTT in CT direction (the GS butterfly equation does not admit a clean pre-twist absorption for forward NWC).'''
    if n < 2 or (n & (n - 1)) != 0:
        raise ValueError(f'n must be a power of 2 and at least 2, got {n}')
    if butterflyType not in ('CT', 'GS'):
        raise ValueError(f"butterflyType must be 'CT' or 'GS', got {butterflyType!r}")
    if negacyclic and butterflyType != 'CT':
        raise ValueError(
            f"negacyclic forward NTT requires butterflyType == 'CT'; "
            f"got {butterflyType!r}. The GS butterfly equation puts the twiddle "
            f"on the subtract branch's bOut, which prevents clean pre-twist absorption "
            f"into per-butterfly twiddles. Use CT for forward NWC and GS for inverse NWC."
        )

    F = GF(modulus)
    if primitiveRoot is None:
        base = F.zeta(2 * n) if negacyclic else F.zeta(n)
    else:
        base = primitiveRoot
    L = int(log2(n))

    twiddles: list[list[int | list[tuple[int, int]]]] = []
    for s in range(L):
        if butterflyType == 'CT':
            stride = 1 << s              # butterflies per group; doubles each stage
            groupSize = stride << 1      # 2^(s+1)
        else:                            # 'GS'
            groupSize = 1 << (L - s)     # halves each stage
            stride = groupSize >> 1
        repeats = n // groupSize         # group pattern repetition count to fill n/2 butterflies
        stepExp = n // groupSize         # exponent step (= n/groupSize)

        if negacyclic:
            # NWC merged twiddle in physical butterfly order: ψ^((2j+1) * stepExp).
            # Derived from path-product analysis: cyclic ω-exponent is j*stepExp; in ψ
            # that's 2j*stepExp; absorbing the per-stage pre-twist factor ψ^(2^(L-1-s)) =
            # ψ^stepExp adds +stepExp, giving (2j+1)*stepExp. Always odd, ranges over
            # the n distinct primitive 2n-th roots that aren't n-th roots.
            baseExps = [(2 * j + 1) * stepExp for j in range(stride)]
        else:
            baseExps = [j * stepExp for j in range(stride)]
        stageExps = baseExps * repeats

        stageTwiddles: list[int | list[tuple[int, int]]] = [int(base ** e) for e in stageExps]
        if useModulusLiftingNaf:
            stageTwiddles = [
                nafTermsModulusLift(t, modulus, maxPower, maxMultipleOfModulus,
                                    maxSearchDepth, beamWidth, maxNumberOfTerms)[1]
                for t in stageTwiddles
            ]
        twiddles.append(stageTwiddles)

    return twiddles


def calculateInttTwiddles(modulus: int,
                          n: int,
                          butterflyType: str,
                          primitiveRoot: int | None = None,
                          negacyclic: bool = False,
                          useModulusLiftingNaf: bool = False,
                          maxPower: int = 95,
                          maxMultipleOfModulus: int = 2**32,
                          maxSearchDepth: int = 3,
                          beamWidth: int = 200,
                          maxNumberOfTerms: int = 96) -> list[list[int | list[tuple[int, int]]]]:
    '''Inverse-NTT twiddle factors. Same shape and conventions as calculateNttTwiddles, but uses w^(-1) (cyclic) or psi^(-1) (negacyclic) as the base so the resulting pipeline computes the inverse transform — modulo the 1/n scaling, which is intentionally omitted; apply externally if needed. For negacyclic, only GS is valid (forward NWC uses CT, inverse NWC uses GS — the standard NWC pairing).'''
    if negacyclic and butterflyType != 'GS':
        raise ValueError(
            f"negacyclic inverse NTT requires butterflyType == 'GS'; "
            f"got {butterflyType!r}. Standard NWC pairing: forward = CT, inverse = GS. "
            f"The CT butterfly equation does not admit a clean post-twist absorption for inverse NWC."
        )
    F = GF(modulus)
    forwardBase = F(primitiveRoot) if primitiveRoot is not None else (F.zeta(2 * n) if negacyclic else F.zeta(n))
    inverseBase = int(forwardBase ** (-1))
    # For negacyclic, calculateNttTwiddles guards against GS — temporarily route through
    # cyclic semantics so we can build the inverse twiddles using the same (2j+1)*stepExp
    # progression but in GS direction. This is symmetric to the CT-forward case.
    if negacyclic:
        # Inverse NWC GS twiddle in physical order: ψ^(-(2j+1) * stepExp) where stepExp uses
        # GS group structure. Build directly here rather than calling calculateNttTwiddles.
        if n < 2 or (n & (n - 1)) != 0:
            raise ValueError(f'n must be a power of 2 and at least 2, got {n}')
        L = int(log2(n))
        psi = F(forwardBase)
        twiddles: list[list[int | list[tuple[int, int]]]] = []
        for s in range(L):
            groupSize = 1 << (L - s)
            stride = groupSize >> 1
            repeats = n // groupSize
            stepExp = n // groupSize
            baseExps = [(2 * j + 1) * stepExp for j in range(stride)]
            stageExps = baseExps * repeats
            stageTwiddles: list[int | list[tuple[int, int]]] = [int(psi ** (-e)) for e in stageExps]
            if useModulusLiftingNaf:
                stageTwiddles = [
                    nafTermsModulusLift(t, modulus, maxPower, maxMultipleOfModulus,
                                        maxSearchDepth, beamWidth, maxNumberOfTerms)[1]
                    for t in stageTwiddles
                ]
            twiddles.append(stageTwiddles)
        return twiddles
    return calculateNttTwiddles(modulus=modulus, n=n, butterflyType=butterflyType,
                                primitiveRoot=inverseBase, negacyclic=False,
                                useModulusLiftingNaf=useModulusLiftingNaf,
                                maxPower=maxPower, maxMultipleOfModulus=maxMultipleOfModulus,
                                maxSearchDepth=maxSearchDepth, beamWidth=beamWidth,
                                maxNumberOfTerms=maxNumberOfTerms)


def loadTwiddlesFromXlsx(path: str, sheetName: str = 'NTT_TWIDDLES') -> list[list[list[tuple[int, int]]]]:
    '''Load twiddles from an xlsx file matching the project's layout. Row 1 holds "Layer 1".."Layer L" headers (an optional trailing label cell like "GS BUTTERFLIES!!!" is ignored). Rows 2..n/2+1 hold per-butterfly twiddles in physical top-to-bottom order, columns are stages. Each cell is either a plain integer or a NAF expression like "-2^91 + 2^43". Returns list[list[list[tuple[int, int]]]] of shape L x (n/2): every twiddle is materialized as a NAF list (matching calculateNttTwiddles output with useModulusLiftingNaf=True). Integer cells are converted via nafTerms.'''
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheetName not in wb.sheetnames:
        raise ValueError(f'sheet {sheetName!r} not found in {path!r}; available sheets: {wb.sheetnames}')
    ws = wb[sheetName]

    L = 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip().lower().startswith('layer'):
            L = c
        else:
            break
    if L == 0:
        raise ValueError(f'no "Layer N" headers found in row 1 of sheet {sheetName!r}')

    halfN = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is None:
            break
        halfN = r - 1

    twiddles: list[list[list[tuple[int, int]]]] = [[[] for _ in range(halfN)] for _ in range(L)]
    for r in range(halfN):
        for layerIdx in range(L):
            v = ws.cell(row=r + 2, column=layerIdx + 1).value
            if isinstance(v, int):
                naf = nafTerms(v)
            elif isinstance(v, str):
                naf = parseNafExpr(v)
            else:
                raise ValueError(f'unexpected cell type at row {r + 2}, column {layerIdx + 1} in sheet {sheetName!r}: {v!r}')
            # canonicalize term order to match calculateNttTwiddles (ascending exponent)
            naf.sort(key=lambda t: t[1])
            twiddles[layerIdx][r] = naf
    return twiddles


def saveTwiddlesToXlsx(twiddles: list[list[int | list[tuple[int, int]]]], path: str, butterflyType: str, sheetName: str = 'NTT_TWIDDLES') -> None:
    '''Save twiddles in the calculated format (output of calculateNttTwiddles) to an xlsx file in the project's layout. Row 1 is "Layer 1".."Layer L" plus a trailing "<TYPE> BUTTERFLIES!!!" label. Rows 2..n/2+1 hold per-butterfly twiddles, written as NAF expression strings (e.g. "-2^91 + 2^43" or "1") so that loadTwiddlesFromXlsx round-trips byte-for-byte. Integer inputs are converted to NAF via nafTerms before writing. If the target file exists the named sheet is replaced and other sheets are preserved; otherwise a new workbook is created.'''
    import openpyxl
    L = len(twiddles)
    if L == 0:
        raise ValueError('twiddles must have at least one stage')
    halfN = len(twiddles[0])
    if any(len(layer) != halfN for layer in twiddles):
        raise ValueError('all stages of twiddles must have the same length')

    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        if sheetName in wb.sheetnames:
            del wb[sheetName]
        ws = wb.create_sheet(sheetName, 0)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheetName

    for layerIdx in range(L):
        ws.cell(row=1, column=layerIdx + 1, value=f'Layer {layerIdx + 1}')
    ws.cell(row=1, column=L + 1, value=f'{butterflyType} BUTTERFLIES!!!')

    for r in range(halfN):
        for layerIdx in range(L):
            v = twiddles[layerIdx][r]
            if isinstance(v, int):
                naf = nafTerms(v)
            elif isinstance(v, list):
                naf = v
            else:
                raise TypeError(f'unexpected twiddle value at stage {layerIdx}, butterfly {r}: {v!r}')
            ws.cell(row=r + 2, column=layerIdx + 1, value=formatNafExpr(naf))

    wb.save(path)

