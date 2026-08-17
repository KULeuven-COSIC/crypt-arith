#!/usr/bin/env python3
"""build_bank.py — bridge operator_modeling and versal_arith.

Generates a bank of N parallel constant multipliers. Designed for
NTT/INTT pre-twist and post-twist banks (``x[i] · ψ^i``-style
arrays) — i.e. cases where the same input fans out to many constant
multipliers in parallel. Per-stage butterfly twiddles do **not** belong
here; those are baked into individual butterfly modules by
``build_butterfly.py``.

Pulls a list of integer constants out of a ``PRE_TWIST`` / ``POST_TWIST``
sheet of ``twiddles.xlsx`` (or a plain integer-per-line file), drops them
into a per-case folder under ``work/``, and invokes the ``versal_arith``
RTL generator (operator ``cmultbank``) so the produced SystemVerilog ends
up in the same case folder.

Layout produced for ``--scenario foo``::

    work/foo/
      foo_constants.txt          # constants fed to the RTL generator
      cmultbank/                 # versal_arith run output
        RTL_generated/           # *.sv (DUT + testbench)
        xdc_generated/           # *.xdc constraints
        testvectors/             # $readmemh inputs / golden outputs
        bitheap_visualization/   # PNGs (if -visualization True)

Run from the project root, e.g.::

    python scripts/build_bank.py \\
        --scenario pre_twist_NTT128 \\
        --sheet PRE_TWIST \\
        --column simple \\
        --width-a 24 \\
        --pipeline-stages 1

If a case file already exists (e.g. ``pre_twist_simpl_numbers.txt``) and you
just want to feed it to the generator, point ``--file`` at it and skip
``--sheet`` / ``--column``::

    python scripts/build_bank.py \\
        --scenario pre_twist_NTT128 \\
        --file work/pre_twist_NTT128/pre_twist_simpl_numbers.txt \\
        --width-a 24 \\
        --modulus 18446744069414584321

Sheets supported by ``--sheet``:

  PRE_TWIST     twiddles.xlsx:    PRETWIST_FACTORS in A; "Most simple bin rep" NAF in B
  POST_TWIST    twiddles.xlsx:    POST TWIST FACTORS in A; simple NAF in B
  PRETWIST      NewTwiddles.xlsx: single raw mod-q column A, no header row
  POSTTWIST     NewTwiddles.xlsx: single raw mod-q column A, header row 1

The two NewTwiddles.xlsx sheets carry **no** pre-lifted NAF column, so they are
raw-only and must be paired with ``--modulus`` to get sparse constants::

    python scripts/build_bank.py \\
        --scenario fourstep64_pretwist \\
        --xlsx NewTwiddles.xlsx --sheet PRETWIST \\
        --width-a 24 --modulus 18446744069414584321

NAF modulus lifting is **optional** and split across two layers:

  * ``--column simple`` (xlsx path): the wrapper itself lifts at extraction
    time by parsing the "Most simple bin rep" NAF expression into an integer.
    The generator then sees already-sparse values; you do **not** need
    ``--modulus``.
  * ``--column raw`` or ``--file <raw integers>``: feed the unlifted mod-q
    residues into the generator. Pass ``--modulus q`` to ask the RTL
    generator to do the lift internally; tune the search with
    ``--lift-max-pow / --lift-max-shift / --lift-depth / --lift-beam``.

If neither lifting path is used, the generator NAF-decomposes each constant
verbatim — fine if your constants are already small, expensive otherwise.
"""

from __future__ import annotations

import argparse
import os
import subprocess
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# Project paths
# ---------------------------------------------------------------------------

PROJECT_ROOT = Path(__file__).resolve().parent.parent
NTT_MODELING_DIR = PROJECT_ROOT / "operator_modeling"
VERSAL_DIR = PROJECT_ROOT / "versal_arith"
DEFAULT_XLSX = PROJECT_ROOT / "twiddles.xlsx"
DEFAULT_WORK_DIR = PROJECT_ROOT / "work"

# ---------------------------------------------------------------------------
# NAF parsing — pulled in from operator_modeling.core.utils so we don't need Sage
# ---------------------------------------------------------------------------

sys.path.insert(0, str(PROJECT_ROOT))
from operator_modeling.core.utils import loadXlsxColumn, parseNafExpr  # noqa: E402


def naf_str_to_int(s: str) -> int:
    """Evaluate a NAF expression string like '-2^91 + 2^43' to an integer."""
    return sum(sign * (1 << k) for sign, k in parseNafExpr(s))


# ---------------------------------------------------------------------------
# Constant extraction
# ---------------------------------------------------------------------------


def constants_from_file(path: Path) -> list[int]:
    out: list[int] = []
    with path.open() as f:
        for lineno, line in enumerate(f, 1):
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            try:
                out.append(int(line))
            except ValueError as e:
                raise ValueError(f"{path}:{lineno}: not an integer: {line!r}") from e
    if not out:
        raise ValueError(f"{path}: no constants read (empty file?)")
    return out


def constants_from_xlsx(
    xlsx: Path,
    sheet: str,
    column: str,
) -> list[int]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit(
            "openpyxl is not installed. Run inside the ntt-sage conda env, or "
            "`pip install openpyxl`."
        ) from e

    wb = load_workbook(xlsx, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(
            f"sheet {sheet!r} not in {xlsx} (have: {wb.sheetnames})"
        )
    ws = wb[sheet]

    if sheet in ("PRE_TWIST", "POST_TWIST"):
        # Row 1 is a header. Cols A=integer, B="Most simple bin rep" NAF,
        # C="Normal bit representation" NAF. A plain header-plus-column read,
        # so it goes through the shared reader in core.utils.
        col_idx = {"raw": 1, "simple": 2}.get(column)
        if col_idx is None:
            raise SystemExit(f"--column must be 'raw' or 'simple' for {sheet}")
        cells = loadXlsxColumn(str(xlsx), sheet, column=col_idx, headerRows=1)
        if column == "raw":
            return [int(cell) for cell in cells]
        return [naf_str_to_int(str(cell)) for cell in cells]

    if sheet in ("PRETWIST", "POSTTWIST"):
        # NewTwiddles.xlsx layout: a single column A of raw mod-q residues,
        # stored as strings, with NO pre-lifted NAF column. Whether row 1 is a
        # header varies per sheet (PRETWIST starts at row 1 with the datum
        # "1" = psi^0; POSTTWIST has a "POST TWIST FACTORS" title), so detect
        # it by asking whether the cell parses as an integer.
        #
        # Because there is no lifted column, these constants MUST be passed to
        # the generator with --modulus so it does the NAF lift internally.
        if column != "raw":
            raise SystemExit(
                f"sheet {sheet!r} has only a single raw column A (no lifted "
                f"'Most simple bin rep'), so --column {column!r} does not "
                f"apply. Re-run with --column raw --modulus <q> so the RTL "
                f"generator performs the NAF lift."
            )
        out: list[int] = []
        for (cell,) in ws.iter_rows(min_row=1, max_col=1, values_only=True):
            if cell is None or str(cell).strip() == "":
                # Trailing blanks end the column; blanks before any data are a
                # layout we don't expect, so stop either way.
                if out:
                    break
                continue
            text = str(cell).strip()
            try:
                out.append(int(text))
            except ValueError:
                if out:
                    raise SystemExit(
                        f"{sheet}: non-integer cell {text!r} after "
                        f"{len(out)} constants"
                    )
                # Header row — skip it and keep going.
                continue
        return out

    raise SystemExit(
        f"sheet {sheet!r} not supported by build_bank.py. Supported sheets: "
        "PRE_TWIST / POST_TWIST (twiddles.xlsx layout) and PRETWIST / "
        "POSTTWIST (NewTwiddles.xlsx layout). Per-stage butterfly twiddles "
        "belong with build_butterfly.py, not in a constant-multiplier bank. "
        "Extract any other constants manually and re-run with --file."
    )


# ---------------------------------------------------------------------------
# Versal RTL generator invocation
# ---------------------------------------------------------------------------


def write_constants_file(constants: list[int], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w") as f:
        for c in constants:
            f.write(f"{c}\n")


def run_versal_cmultbank(
    constants_path: Path,
    output_dir: Path,
    width_a: int,
    pipeline_stages: int,
    signed_input: bool,
    test_size: int,
    visualization: bool,
    modulus: int | None,
    lift_max_pow: int,
    lift_max_shift: int,
    lift_depth: int,
    lift_beam: int,
) -> None:
    """Subprocess-call versal_arith/cli.py -operator cmultbank with cwd inside
    versal_arith/.
    """
    cmd = [
        sys.executable,
        "cli.py",
        "-operator", "cmultbank",
        "-txt_file_name", str(constants_path.resolve()),
        "-width_a", str(width_a),
        "-pipeline_stages", str(pipeline_stages),
        "-output_dir", str(output_dir.resolve()),
        "-signed_input", "True" if signed_input else "False",
        "-test_size", str(test_size),
        "-visualization", "True" if visualization else "False",
    ]
    if modulus is not None:
        cmd += [
            "-modulus", str(modulus),
            "-lift_max_pow", str(lift_max_pow),
            "-lift_max_shift", str(lift_max_shift),
            "-lift_depth", str(lift_depth),
            "-lift_beam", str(lift_beam),
        ]
    print(f"[build_bank] running: {' '.join(cmd)}")
    print(f"[build_bank] cwd:     {VERSAL_DIR}")
    subprocess.run(cmd, cwd=VERSAL_DIR, check=True)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    p = argparse.ArgumentParser(
        formatter_class=argparse.RawDescriptionHelpFormatter,
        description=__doc__,
    )

    # Where to put things
    p.add_argument("--scenario", required=True,
                   help="name of the work-dir subfolder (e.g. pre_twist_NTT128)")
    p.add_argument("--work-dir", default=str(DEFAULT_WORK_DIR),
                   help=f"root work directory (default: {DEFAULT_WORK_DIR})")

    # Where the constants come from
    src = p.add_mutually_exclusive_group()
    src.add_argument("--sheet",
                     choices=("PRE_TWIST", "POST_TWIST", "PRETWIST", "POSTTWIST"),
                     help="xlsx sheet name. PRE_TWIST / POST_TWIST use the "
                          "twiddles.xlsx layout (header row, lifted NAF in "
                          "column B); PRETWIST / POSTTWIST use the "
                          "NewTwiddles.xlsx layout (single raw column A)")
    src.add_argument("--file",
                     help="path to a plain integer-per-line file (skips xlsx)")
    p.add_argument("--xlsx", default=str(DEFAULT_XLSX),
                   help=f"path to the twiddle workbook (default: {DEFAULT_XLSX})")
    p.add_argument("--column", choices=("simple", "raw"), default=None,
                   help="for PRE_TWIST/POST_TWIST: 'simple' = column B "
                        "(already-lifted, sparse NAF — the default for those "
                        "sheets), 'raw' = column A (original mod-q integer; "
                        "pair with --modulus to lift inside the RTL "
                        "generator). PRETWIST/POSTTWIST are raw-only and "
                        "default accordingly.")

    # Versal generator parameters (forwarded to cli.py)
    p.add_argument("--width-a", type=int, default=24,
                   help="input bit-width A (default: 24)")
    p.add_argument("--pipeline-stages", type=int, default=1,
                   help="pipeline stages (default: 1)")
    p.add_argument("--signed-input", action="store_true",
                   help="treat A as signed two's complement")
    p.add_argument("--test-size", type=int, default=1000,
                   help="random test vectors (default: 1000)")
    p.add_argument("--no-visualization", action="store_true",
                   help="skip bit-heap PNGs (faster, no matplotlib needed)")
    p.add_argument("--dry-run", action="store_true",
                   help="extract and write constants only; skip RTL generation")

    # NAF modulus-lift control (forwarded only when --modulus is set)
    p.add_argument("--modulus", type=int, default=None,
                   help="modulus q for sparser-NAF lifting inside the RTL "
                        "generator. Pass this when feeding raw mod-q residues "
                        "(e.g. PRE_TWIST column A) so the generator lifts each "
                        "constant. Skip it if your constants are already "
                        "lifted or otherwise small. Goldilocks q = "
                        "18446744069414584321.")
    p.add_argument("--lift-max-pow", type=int, default=96,
                   help="cap on NAF exponents in the lifted form (default: 96). "
                        "Only consulted when --modulus is set.")
    p.add_argument("--lift-max-shift", type=int, default=32,
                   help="max shift t such that q*2^t may enter the search "
                        "frontier (default: 32). Only consulted when --modulus "
                        "is set.")
    p.add_argument("--lift-depth", type=int, default=3,
                   help="beam-search iterations (default: 3). Only consulted "
                        "when --modulus is set.")
    p.add_argument("--lift-beam", type=int, default=200,
                   help="beam width (default: 200). Only consulted when "
                        "--modulus is set.")

    # Remote simulation control
    p.add_argument("--remote-sim", action="store_true",
                   help="after generation, stage the run on the V80 server "
                        "and run Vivado batch simulation; pull artifacts to "
                        "<work_dir>/<scenario>/sim_remote/")
    p.add_argument("--remote-server",
                   default=os.environ.get("V80_SERVER", "v80-server"),
                   help="SSH alias of the V80 server (default: $V80_SERVER or 'v80-server'; "
                        "configure in ~/.ssh/config)")
    p.add_argument("--remote-root",
                   default=os.environ.get("V80_REMOTE_ROOT", "~/AMD_V80_dev"),
                   help="path to the Vivado project on the server "
                        "(default: $V80_REMOTE_ROOT or '~/AMD_V80_dev')")

    args = p.parse_args()

    work_dir = Path(args.work_dir).resolve()
    case_dir = work_dir / args.scenario
    case_dir.mkdir(parents=True, exist_ok=True)
    constants_path = case_dir / f"{args.scenario}_constants.txt"

    # ----- 1. Acquire the constants -----
    if args.file:
        src_path = Path(args.file).resolve()
        if not src_path.is_file():
            raise SystemExit(f"--file not found: {src_path}")
        constants = constants_from_file(src_path)
        print(f"[build_bank] loaded {len(constants)} constants from {src_path}")
    elif args.sheet:
        xlsx_path = Path(args.xlsx).resolve()
        if not xlsx_path.is_file():
            raise SystemExit(f"--xlsx not found: {xlsx_path}")
        # Per-sheet default: the twiddles.xlsx sheets carry a pre-lifted NAF in
        # column B, the NewTwiddles.xlsx ones only a raw residue in column A.
        column = args.column
        if column is None:
            column = "raw" if args.sheet in ("PRETWIST", "POSTTWIST") else "simple"
        constants = constants_from_xlsx(
            xlsx_path, args.sheet, column
        )
        print(f"[build_bank] extracted {len(constants)} constants from "
              f"{xlsx_path}!{args.sheet} (column={column})")
        if column == "raw" and args.modulus is None:
            print("[build_bank] WARNING: raw mod-q residues with no --modulus; "
                  "the generator will NAF-decompose them verbatim (dense, "
                  "expensive). Pass --modulus 18446744069414584321 to lift.")
    else:
        raise SystemExit("must pass either --sheet or --file")

    # ----- 2. Snapshot them inside the case folder -----
    write_constants_file(constants, constants_path)
    print(f"[build_bank] wrote {constants_path}")

    if args.dry_run:
        print("[build_bank] --dry-run: skipping RTL generation")
        return

    # ----- 3. Hand off to versal_arith -----
    run_versal_cmultbank(
        constants_path=constants_path,
        output_dir=case_dir,
        width_a=args.width_a,
        pipeline_stages=args.pipeline_stages,
        signed_input=args.signed_input,
        test_size=args.test_size,
        visualization=not args.no_visualization,
        modulus=args.modulus,
        lift_max_pow=args.lift_max_pow,
        lift_max_shift=args.lift_max_shift,
        lift_depth=args.lift_depth,
        lift_beam=args.lift_beam,
    )

    # ----- 4. Print where things landed -----
    bank_dir = case_dir / "cmultbank"
    print(f"[build_bank] done.")
    print(f"  constants: {constants_path}")
    print(f"  RTL:       {bank_dir / 'RTL_generated'}")
    print(f"  XDC:       {bank_dir / 'xdc_generated'}")
    print(f"  TBs/data:  {bank_dir / 'testvectors'}")

    # ----- 5. Optional: stage and simulate on the V80 server -----
    if args.remote_sim:
        print(f"[build_bank] kicking off remote sim on {args.remote_server}")
        rc = subprocess.run([
            sys.executable,
            str(PROJECT_ROOT / "scripts" / "run_remote_sim.py"),
            "--run-dir",     str(bank_dir),
            "--server",      args.remote_server,
            "--remote-root", args.remote_root,
            "--pull-to",     str(case_dir / "sim_remote"),
        ]).returncode
        if rc != 0:
            raise SystemExit(f"[build_bank] remote sim returned exit {rc}")


if __name__ == "__main__":
    main()
